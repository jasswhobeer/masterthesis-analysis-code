import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import matplotlib.ticker as ticker
from matplotlib import cm
from matplotlib.patches import Patch
from scipy.optimize import curve_fit
from scipy import stats
from scipy.stats import gaussian_kde
from scipy.ndimage import gaussian_filter1d
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path
import warnings
warnings.filterwarnings('ignore')

INPUT_DIR = Path(__file__).parent
FILES = {
    'P1_CONF':  'Fibrinogen Analyse Plate 1 Confluence.txt',
    'P1_FOCUS': 'Fibrinogen Analyse Plate 1 Focus Position.txt',
    'P2_CONF':  'Fibrinogen Analyse Plate 2 Confluence.txt',
    'P2_FOCUS': 'Fibrinogen Analyse Plate 2 Focus Position.txt',
}

KDE_BW          = 50.0
FOCUS_TOL       = 40.0
CELL_FOCUS_MIN  = 3000.0

MIN_VALID_PTS   = 8
MIN_TIME_SPAN   = 40.0

T_GRID_N        = 500
T_REF           = 140.0

EXPLICIT_EXCLUSIONS = {
    ('P1', 'Fibrinogen Coating (Dried) 100 mg/ml (A5)'),
    ('P1', 'Fibrinogen Coating 100 mg/ml (B4)'),
}

TOST_BOUND_A    = 10.0
TOST_BOUND_MU   = 0.30

WELLS_PER_CONDITION = {
    'Dried Fibrinogen': {
        'P1': ['Fibrinogen Coating (Dried) 100 mg/ml (A1)',
               'Fibrinogen Coating (Dried) 100 mg/ml (A2)',
               'Fibrinogen Coating (Dried) 100 mg/ml (A3)',
               'Fibrinogen Coating (Dried) 100 mg/ml (A4)',
               'Fibrinogen Coating (Dried) 100 mg/ml (A5)'],
        'P2': ['Fibrinogen Coating (Dried) 100 mg/ml (A1)',
               'Fibrinogen Coating (Dried) 100 mg/ml (A2)',
               'Fibrinogen Coating (Dried) 100 mg/ml (A3)']},
    'Wet Fibrinogen': {
        'P1': ['Fibrinogen Coating 100 mg/ml (B1)',
               'Fibrinogen Coating 100 mg/ml (B2)',
               'Fibrinogen Coating 100 mg/ml (B3)',
               'Fibrinogen Coating 100 mg/ml (B4)',
               'Fibrinogen Coating 100 mg/ml (B5)'],
        'P2': ['Fibrinogen Coating 100 mg/ml (B1)',
               'Fibrinogen Coating 100 mg/ml (B2)',
               'Fibrinogen Coating 100 mg/ml (B3)']},
    'COC Untreated': {
        'P1': ['COC untreated (C1)', 'COC untreated (C2)',
               'COC untreated (C3)', 'COC untreated (C4)',
               'COC untreated (C5)'],
        'P2': ['COC untreated (C1)', 'COC untreated (C2)',
               'COC untreated (C3)']},
    'Well Only': {
        'P1': ['Well only Reference (A6)', 'Well only Reference (B6)',
               'Well only Reference (C6)', 'Well only Reference (D6)'],
        'P2': ['Well only Reference (D1)', 'Well only Reference (D2)',
               'Well only Reference (D3)']},
    'Insert Only': {
        'P1': ['Inlet only Reference (D1)', 'Inlet only Reference (D2)',
               'Inlet only Reference (D3)', 'Inlet only Reference (D4)',
               'Inlet only Reference (D5)'],
        'P2': []},
}
COND_LIST = list(WELLS_PER_CONDITION.keys())
COLORS = {c: cm.viridis(v) for c, v in
          zip(COND_LIST, [0.05, 0.28, 0.52, 0.75, 0.92])}

def load_incucyte(path):
    with open(path, encoding='utf-8') as f:
        lines = f.readlines()
    h = next(i for i, l in enumerate(lines) if 'Elapsed' in l)
    df = pd.read_csv(path, sep='\t', decimal=',', skiprows=h, encoding='utf-8')
    df.columns = df.columns.str.strip()
    df['Elapsed'] = pd.to_numeric(df['Elapsed'], errors='coerce')
    return df.dropna(subset=['Elapsed'])

def find_cell_plane(f_values):
    fv = f_values[(~np.isnan(f_values)) & (f_values >= 0)]
    if len(fv) < 5:
        return None, 0.0
    if fv.std() < 1.0:
        pos = float(fv.mean())
        return (pos, 1.0) if pos >= CELL_FOCUS_MIN else (None, 0.0)
    bw = KDE_BW / fv.std()
    kde = gaussian_kde(fv, bw_method=bw)
    x = np.linspace(fv.min() - 50, fv.max() + 50, 2000)
    d = kde(x)
    peaks = [x[i] for i in range(1, len(d)-1)
             if d[i] > d[i-1] and d[i] > d[i+1]]
    if not peaks:
        pos = float(fv.mean())
        return (pos, 1.0) if pos >= CELL_FOCUS_MIN else (None, 0.0)
    high = [(p, np.sum(np.abs(fv - p) <= FOCUS_TOL) / len(fv))
            for p in peaks if p >= CELL_FOCUS_MIN]
    if not high:
        return None, 0.0
    return max(high, key=lambda x: x[1])

def sawtooth_index(t, c):
    if len(c) < 6:
        return np.nan
    rng = c.max() - c.min()
    if rng < 1e-6:
        return np.nan
    raw = np.mean(np.abs(np.diff(c)))
    smooth = np.mean(np.abs(np.diff(gaussian_filter1d(c, sigma=2.0))))
    return (raw - smooth) / rng

def pre_peak_rho(t, c):
    if len(t) < 4:
        return np.nan
    pk = max(int(np.argmax(c)), 3)
    r, _ = stats.spearmanr(t[:pk+1], c[:pk+1])
    return float(r) if not np.isnan(r) else np.nan

def gompertz_offset(t, C0, A, mu, lag):
    return C0 + A * np.exp(-np.exp((mu * np.e / A) * (lag - t) + 1.0))

def gompertz_offset_decline(t, C0, A, mu, lag, kd, td):
    g = gompertz_offset(t, C0, A, mu, lag)
    s = 1.0 / (1.0 + np.exp(-(t - td) / 2.0))
    return g * (1 - s) + g * s * np.exp(-kd * np.clip(t - td, 0, None))

def logistic4_offset(t, C0, A, k, tmid):
    return C0 + (A - C0) / (1.0 + np.exp(-k * (t - tmid)))

def linear_fn(t, a, b):
    return a * t + b

def _fit(fn, t, y, p0, bounds, k):
    try:
        popt, _ = curve_fit(fn, t, y, p0=p0, bounds=bounds, maxfev=40000)
        yp = fn(t, *popt)
        ss_res = np.sum((y - yp)**2)
        ss_tot = np.sum((y - y.mean())**2)
        r2  = 1 - ss_res / ss_tot if ss_tot > 1e-10 else np.nan
        n   = len(t)
        K   = k + 1
        aic = n * np.log(max(ss_res / n, 1e-12)) + 2 * K
        nenner = n - K - 1
        aicc = aic + (2 * K * (K + 1) / nenner) if nenner > 0 else np.inf
        return {'popt': popt, 'r2': float(r2), 'aic': float(aic),
                'aicc': float(aicc), 'fn': fn, 'k': k}
    except Exception:
        return None

def fit_well(t, y):
    n       = len(t)
    C0_init = max(float(y[0]), 0.5)
    candidates = {}

    A0 = max(float(y.max()) - C0_init, 1.0)
    r = _fit(gompertz_offset, t, y,
             p0=[C0_init, A0, 0.5, max(t[0], 1.0)],
             bounds=([0.0, 0.1, 0.001, 0.0],
                     [100.0, 100.0, 20.0, t.max()*0.8]),
             k=4)
    if r: candidates['gompertz'] = r

    pk  = int(np.argmax(y))
    td0 = t[pk] if pk > n//3 else t[n//2]
    kd0 = max(-np.log(max(y[-1]/max(y[pk], 0.1), 0.01)) /
              max(t[-1]-t[pk], 1.0), 0.001) if pk < n-3 else 0.005
    A_cap = max(float(y.max()) * 1.5, 5.0)
    r = _fit(gompertz_offset_decline, t, y,
             p0=[C0_init, max(float(y[pk])-C0_init, 1.0),
                 0.5, max(t[0], 1.0), kd0, td0],
             bounds=([0.0, 0.1, 0.001, 0.0, 0.0, t[min(3, n-1)]],
                     [100.0, min(A_cap, 100.0), 20.0,
                      t.max()*0.7, 0.3, t.max()]),
             k=6)
    if r: candidates['gompertz_decline'] = r

    A0 = min(max(float(y.max()) * 1.1, C0_init + 1.0), 100.0)
    r = _fit(logistic4_offset, t, y,
             p0=[C0_init, A0, 0.05, float(np.median(t))],
             bounds=([0.0, C0_init, 0.001, t.min()-20],
                     [100.0, 100.0, 5.0, t.max()+20]),
             k=4)
    if r: candidates['logistic4'] = r

    r = _fit(linear_fn, t, y,
             p0=[0.0, float(y.mean())],
             bounds=([-1.0, -10.0], [1.0, 110.0]),
             k=2)
    if r: candidates['linear'] = r

    if not candidates:
        return None

    best_aicc = min(v['aicc'] for v in candidates.values())
    eligible = sorted(
        [(name, v) for name, v in candidates.items()
         if v['aicc'] - best_aicc < 4.0],
        key=lambda x: (x[1]['k'], x[1]['aicc'])
    )
    name, chosen = eligible[0]

    p = chosen['popt']
    if name == 'gompertz':
        C0, zuwachs, mu_max, lag = p[0], p[1], p[2], p[3]
        kd, td = 0.0, t.max()
    elif name == 'gompertz_decline':
        C0, zuwachs, mu_max, lag, kd, td = p[0], p[1], p[2], p[3], p[4], p[5]
    elif name == 'logistic4':
        C0, plateau_4pl = p[0], p[1]
        zuwachs = plateau_4pl - C0
        mu_max  = p[2] * zuwachs / 4.0
        lag     = max(p[3] - 2.0 / max(p[2], 1e-6), 0.0)
        kd, td  = 0.0, t.max()
    else:
        C0      = max(float(p[1]), 0.0)
        zuwachs = max(float(p[0]) * T_REF, 0.0)
        mu_max  = max(float(p[0]), 0.0)
        lag, kd, td = 0.0, 0.0, t.max()

    A = C0 + zuwachs

    chosen.update({
        'name': name, 'C0': C0, 'A': A, 'A_gain': zuwachs,
        'mu_max': mu_max, 'lag': lag, 'kd': kd, 'td': td,
        'aicc_all': {k: v['aicc'] for k, v in candidates.items()},
    })
    return chosen

def assess_well(well, plate_label, conf_df, focus_df):
    r = {
        'well': well, 'plate': plate_label,
        'pass': False, 'reason': '',
        'n_total': 0, 'n_valid': 0, 'valid_fraction': 0.0,
        'cell_plane_um': np.nan, 'cell_plane_fraction': 0.0,
        'intra_focus_sd': np.nan, 'sawtooth': np.nan, 'rho': np.nan,
        't_all': None, 'c_all': None,
        't_valid': None, 'c_valid': None,
        'fit': None, 'fit_name': '',
        'C0': np.nan, 'A': np.nan, 'A_gain': np.nan,
        'mu_max': np.nan, 'lag': np.nan,
        'kd': np.nan, 'td': np.nan,
        'fit_r2': np.nan, 'obs_max': np.nan,
    }
    if well not in conf_df.columns or well not in focus_df.columns:
        r['reason'] = 'column missing'; return r

    if (plate_label, well) in EXPLICIT_EXCLUSIONS:
        r['reason'] = 'explicit exclusion (documented, see script header)'
        r['t_valid'] = np.array([])
        r['c_valid'] = np.array([])
        r['fit'] = None
        return r

    merged = pd.merge(
        conf_df[['Elapsed', well]].rename(columns={well: '_c'}),
        focus_df[['Elapsed', well]].rename(columns={well: '_f'}),
        on='Elapsed', how='inner'
    ).dropna(subset=['Elapsed'])
    t = merged['Elapsed'].values.astype(float)
    c = merged['_c'].values.astype(float)
    f = merged['_f'].values.astype(float)
    r['n_total'] = len(t)
    r['t_all'] = t; r['c_all'] = c

    valid_raw = (~np.isnan(c)) & (~np.isnan(f)) & (f >= 0)

    pos, frac = find_cell_plane(f[valid_raw])
    if pos is None:
        r['reason'] = 'no cell-plane mode ≥ 3000 µm'; return r
    r['cell_plane_um'] = pos
    r['cell_plane_fraction'] = frac

    in_band = valid_raw & (np.abs(f - pos) <= FOCUS_TOL)
    n_val = int(in_band.sum())
    r['n_valid'] = n_val
    r['valid_fraction'] = n_val / len(t)

    if n_val < MIN_VALID_PTS:
        r['reason'] = f'only {n_val} valid pts < {MIN_VALID_PTS}'; return r

    idx = np.argsort(t[in_band])
    tv = t[in_band][idx]; cv = c[in_band][idx]
    tspan = tv[-1] - tv[0]

    if tspan < MIN_TIME_SPAN:
        r['reason'] = f'time span {tspan:.0f} h < {MIN_TIME_SPAN} h'; return r

    r['t_valid'] = tv; r['c_valid'] = cv
    r['obs_max'] = float(cv.max())
    r['intra_focus_sd'] = float(np.std(f[in_band], ddof=1)) if n_val > 1 else 0.0
    r['sawtooth'] = sawtooth_index(tv, cv)
    r['rho'] = pre_peak_rho(tv, cv)

    fit = fit_well(tv, cv)
    if fit is None:
        r['reason'] = 'no model converged'; return r

    r['fit'] = fit
    r['fit_name'] = fit['name']
    r['C0']     = fit.get('C0', np.nan)
    r['A']      = fit['A']
    r['A_gain'] = fit['A_gain']
    r['mu_max'] = fit['mu_max']
    r['lag']    = fit['lag']
    r['kd']     = fit['kd']
    r['td']     = fit['td']
    r['fit_r2'] = fit['r2']
    r['pass']   = True
    r['reason'] = 'passed'
    return r

DATA = {
    'P1': (load_incucyte(INPUT_DIR / FILES['P1_CONF']),
           load_incucyte(INPUT_DIR / FILES['P1_FOCUS'])),
    'P2': (load_incucyte(INPUT_DIR / FILES['P2_CONF']),
           load_incucyte(INPUT_DIR / FILES['P2_FOCUS'])),
}

print('='*65)
print('  Fibrinogen v3 — minimaler Fokus-Filter, Fit-basierte Plots')
print('='*65)

all_wells = []
for cond, plates in WELLS_PER_CONDITION.items():
    for plate, wells in plates.items():
        cdf, fdf = DATA[plate]
        for well in wells:
            r = assess_well(well, plate, cdf, fdf)
            r['condition'] = cond
            all_wells.append(r)

print('\nErgebnis pro Condition:')
for cond in COND_LIST:
    in_c  = [r for r in all_wells if r['condition'] == cond]
    passed = [r for r in in_c if r['pass']]
    print(f'  {cond:<26} {len(passed):2d}/{len(in_c):2d} passed')

print('\nDetail (nur excluded):')
for r in all_wells:
    if not r['pass']:
        print(f'  [{r["plate"]}] {r["well"][-10:]:<12} -> {r["reason"]}')

def cohens_d(x, y):
    if len(x) < 2 or len(y) < 2: return np.nan
    sp = np.sqrt(((len(x)-1)*np.var(x,ddof=1) +
                  (len(y)-1)*np.var(y,ddof=1)) / (len(x)+len(y)-2))
    return (np.mean(x)-np.mean(y))/sp if sp > 0 else np.nan

def tost(x, y, bound):
    nx, ny = len(x), len(y)
    if nx < 2 or ny < 2: return np.nan, False
    diff = np.mean(x) - np.mean(y)
    vx, vy = np.var(x, ddof=1) / nx, np.var(y, ddof=1) / ny
    se = np.sqrt(vx + vy)
    if se == 0: return np.nan, False
    df = (vx + vy)**2 / (vx**2 / (nx - 1) + vy**2 / (ny - 1))
    p = max(1-stats.t.cdf((diff-(-bound))/se, df),
            stats.t.cdf((diff-bound)/se, df))
    return float(p), bool(p < 0.05)

def pairwise_mwu(all_wells, param):
    vals = {}
    for r in all_wells:
        if r['pass'] and not np.isnan(r.get(param, np.nan)):
            vals.setdefault(r['condition'], []).append(r[param])
    conds = list(vals.keys())
    pairs = [(a,b) for i,a in enumerate(conds) for b in conds[i+1:]]
    n_p = max(len(pairs), 1)
    rows = []
    for a, b in pairs:
        xa, xb = np.array(vals[a]), np.array(vals[b])
        if len(xa) < 2 or len(xb) < 2:
            rows.append({'A':a,'B':b,'n_A':len(xa),'n_B':len(xb),
                         'mean_A':np.nan,'mean_B':np.nan,
                         'U':np.nan,'p':np.nan,'p_bonf':np.nan,'d':np.nan})
            continue
        U, p = stats.mannwhitneyu(xa, xb, alternative='two-sided')
        rows.append({'A':a,'B':b,'n_A':len(xa),'n_B':len(xb),
                     'mean_A':float(np.mean(xa)),'mean_B':float(np.mean(xb)),
                     'U':float(U),'p':float(p),
                     'p_bonf':float(min(p*n_p,1.0)),
                     'd':float(cohens_d(xa,xb))})
    return pd.DataFrame(rows)

def plot_qc(all_wells, fname='fibrinogen_v3_qc.png'):
    n = len(all_wells)
    nc = 6; nr = int(np.ceil(n/nc))
    fig, axes = plt.subplots(nr, nc, figsize=(nc*2.8, nr*2.4),
                             constrained_layout=True)
    axes = np.atleast_2d(axes).flatten()
    t_grid = np.linspace(0, 140, 400)
    for ax, r in zip(axes, all_wells):
        if r['t_all'] is None:
            ax.axis('off'); continue
        ax.scatter(r['t_all'], r['c_all'], s=5, c='lightgray',
                   alpha=0.5, linewidths=0, zorder=1)
        color = '#1a4a2e' if r['pass'] else '#a92020'
        if r['pass'] and r['fit'] is not None:
            ax.scatter(r['t_valid'], r['c_valid'], s=7, c=color,
                       alpha=0.85, linewidths=0, zorder=3)
            yf = r['fit']['fn'](t_grid, *r['fit']['popt'])
            ax.plot(t_grid, np.clip(yf, 0, 105), '-', c='#e54f1f',
                    lw=1.5, zorder=4)
        short = (r['well']
                 .replace('Fibrinogen Coating','FC')
                 .replace('100 mg/ml','')
                 .replace('only Reference','ref').strip())[:22]
        label = 'PASS' if r['pass'] else 'EXCL'
        ax.set_title(f'{r["plate"]} {short}\n[{label}]',
                     fontsize=7, color=color)
        ax.set_xlabel('t (h)', fontsize=6)
        ax.set_ylabel('Conf %', fontsize=6)
        for sp in ax.spines.values():
            sp.set_visible(True)
        ax.tick_params(labelsize=6, direction='in', top=False, right=False)
        ax.grid(False)
        ax.set_xlim(0, 142); ax.set_ylim(-2, 105)
        if not r['pass']:
            ax.text(0.5, 0.5, 'EXCL', transform=ax.transAxes,
                    ha='center', va='center', fontsize=12,
                    color='#a92020', alpha=0.3, fontweight='bold')
    for ax in axes[n:]:
        ax.axis('off')
    fig.suptitle('Per-Well QC: grau=alle Rohdaten, grün=fokus-valide Punkte, '
                 'orange=Fit ab t=0', fontsize=20)
    fig.savefig(fname, dpi=180, bbox_inches='tight', facecolor='white')
    plt.close(fig)
    print(f'  QC:       {fname}')

def plot_growth(all_wells, fname='fibrinogen_v3_growth.png'):
    fig, ax = plt.subplots(figsize=(13, 8))
    fig.patch.set_facecolor('white'); ax.set_facecolor('white')
    t_grid = np.linspace(0, 140, T_GRID_N)
    legend_handles = []

    for cond in COND_LIST:
        col = COLORS[cond]
        passed = [r for r in all_wells
                  if r['condition'] == cond and r['pass']]
        n_p = len(passed)
        n_f = sum(1 for r in all_wells
                  if r['condition'] == cond and not r['pass'])

        for r in passed:
            yf = np.clip(r['fit']['fn'](t_grid, *r['fit']['popt']), 0, 100)
            ax.plot(t_grid, yf, '-',
                    color=col, alpha=0.22, lw=1.0, zorder=2)

        if n_p == 0:
            legend_handles.append(
                Patch(facecolor=col, edgecolor='black',
                      label=f'{cond}  (n={n_p})'))
            continue

        fit_matrix = np.array([
            np.clip(r['fit']['fn'](t_grid, *r['fit']['popt']), 0, 100)
            for r in passed
        ])
        mean_curve = fit_matrix.mean(axis=0)
        if n_p > 1:
            se_curve = fit_matrix.std(axis=0, ddof=1) / np.sqrt(n_p)
            ci_lo = mean_curve - 1.96 * se_curve
            ci_hi = mean_curve + 1.96 * se_curve
        else:
            ci_lo = ci_hi = mean_curve

        ax.plot(t_grid, mean_curve, '-', color=col, lw=2.8, zorder=5)
        ax.fill_between(t_grid, ci_lo, ci_hi,
                        color=col, alpha=0.20, zorder=4)

        legend_handles.append(
            Patch(facecolor=col, edgecolor='black',
                  label=f'{cond}  (n={n_p})'))

    for sp in ax.spines.values():
        sp.set_visible(True); sp.set_linewidth(1.5); sp.set_color('black')
    ax.tick_params(direction='in', length=6, width=1.4,
                   top=False, right=False, labelsize=12)
    ax.grid(False)
    ax.set_xlim(0, 142)
    ax.set_ylim(0, None)
    ax.set_xlabel('Elapsed Time (h)', fontsize=16, labelpad=10)
    ax.set_ylabel('Phase Object Confluence (%)', fontsize=16, labelpad=10)
    ax.set_title('HUVEC Cell Confluence on Fibrinogen-Coated COC Surfaces',
                 fontsize=20, pad=12)
    leg = ax.legend(handles=legend_handles, loc='upper left',
                    fontsize=14, frameon=True, fancybox=False,
                    edgecolor='black', facecolor='white', framealpha=0.95)
    leg.get_frame().set_linewidth(1.2)
    plt.tight_layout()
    plt.savefig(fname, dpi=300, bbox_inches='tight', facecolor='white')
    plt.close(fig)
    print(f'  Growth:   {fname}')

def plot_passrate(all_wells, fname='fibrinogen_v3_passrate.png'):
    fig, ax = plt.subplots(figsize=(10, 5))
    labels = [c.replace(' (TC-Plastik)','').replace(' (kein COC)','')
              for c in COND_LIST]
    n_pass = [sum(1 for r in all_wells
                  if r['condition']==c and r['pass']) for c in COND_LIST]
    n_fail = [sum(1 for r in all_wells
                  if r['condition']==c and not r['pass']) for c in COND_LIST]
    x = np.arange(len(COND_LIST))
    ax.bar(x, n_pass, color='#1a4a2e', edgecolor='black',
           label='Passing (fit)')
    ax.bar(x, n_fail, bottom=n_pass, color='#a92020', edgecolor='black',
           alpha=0.7, label='Excluded')
    for i, (g, f) in enumerate(zip(n_pass, n_fail)):
        ax.text(i, g+f+0.15, f'n={g+f}', ha='center',
                fontsize=10, fontweight='bold')
    ax.set_xticks(x); ax.set_xticklabels(labels, rotation=18, ha='right')
    ax.set_ylabel('Number of Wells', fontsize=12)
    ax.set_title('Well Inclusion per Condition', fontsize=12,
                 fontweight='bold')
    ax.legend(frameon=True, edgecolor='black')
    for sp in ax.spines.values(): sp.set_linewidth(1.2)
    ax.tick_params(direction='in', length=5, width=1.0,
                   top=True, right=True)
    plt.tight_layout()
    plt.savefig(fname, dpi=300, bbox_inches='tight', facecolor='white')
    plt.close(fig)
    print(f'  Passrate: {fname}')

def plot_params(all_wells, fname='fibrinogen_v3_params.png'):
    metrics = [
        ('A',      'Plateau A (%)',          'fitted upper asymptote'),
        ('mu_max', 'Growth Rate µ_max (%/h)','max. specific growth rate'),
        ('lag',    'Lag Time λ (h)',         'lag phase'),
        ('obs_max','Observed Max. (%)',       'highest measured value'),
    ]
    fig, axes = plt.subplots(1, 4, figsize=(16, 5))
    fig.patch.set_facecolor('white')
    jitter_rng = np.random.default_rng(20260727)
    for ax, (key, ylabel, subtitle) in zip(axes, metrics):
        data, labels, colors = [], [], []
        for cond in COND_LIST:
            vs = [r[key] for r in all_wells
                  if r['condition']==cond and r['pass']
                  and not np.isnan(r[key])]
            if vs:
                data.append(vs)
                labels.append(cond.replace(' (TC-Plastik)','')
                              .replace(' (kein COC)',''))
                colors.append(COLORS[cond])
        if not data:
            ax.text(0.5,0.5,'no data',transform=ax.transAxes,
                    ha='center',va='center'); continue
        bp = ax.boxplot(data, patch_artist=True, widths=0.55,
                        medianprops=dict(color='black',linewidth=1.5),
                        showfliers=False)
        for patch, c in zip(bp['boxes'], colors):
            patch.set_facecolor(c); patch.set_alpha(0.55)
            patch.set_edgecolor('black'); patch.set_linewidth(1.0)
        for i, (vs, c) in enumerate(zip(data, colors), 1):
            ax.scatter(jitter_rng.normal(i, 0.06, len(vs)), vs,
                       color=c, edgecolor='black', s=46,
                       zorder=3, linewidth=0.6, alpha=0.85)
        for sp in ax.spines.values(): sp.set_linewidth(1.2)
        ax.tick_params(direction='in', length=5, width=1.0,
                       top=True, right=True, labelsize=9)
        ax.set_ylabel(ylabel, fontsize=11)
        ax.set_title(subtitle, fontsize=10)
        ax.set_xticklabels(labels, rotation=20, ha='right', fontsize=9)
        ax.yaxis.set_minor_locator(ticker.AutoMinorLocator(2))
    fig.suptitle('Per-Well Growth Parameters (all passing wells)',
                 fontsize=13, fontweight='bold')
    plt.tight_layout()
    plt.savefig(fname, dpi=300, bbox_inches='tight', facecolor='white')
    plt.close(fig)
    print(f'  Params:   {fname}')

plot_qc(all_wells)
plot_growth(all_wells)
plot_passrate(all_wells)
plot_params(all_wells)

print('\n--- Paarweise MWU + Bonferroni (A, mu_max) ---')
for param in ['A', 'mu_max']:
    print(f'\n{param}:')
    df = pairwise_mwu(all_wells, param)
    if not df.empty:
        print(df[['A','B','n_A','n_B','mean_A','mean_B',
                  'p_bonf','d']].to_string(index=False))

def gv(cond, key):
    return [r[key] for r in all_wells
            if r['condition']==cond and r['pass']
            and not np.isnan(r[key])]

print('\n--- TOST Äquivalenz Dried vs. Wet ---')
for key, bound, unit in [('A', TOST_BOUND_A, 'pp'),
                         ('mu_max', None, 'rel')]:
    x = gv('Dried Fibrinogen', key)
    y = gv('Wet Fibrinogen', key)
    if not x or not y: continue
    b = bound if bound else TOST_BOUND_MU * np.mean(y)
    p, eq = tost(x, y, b)
    print(f'  {key}: dried={np.mean(x):.3f}(n={len(x)}), '
          f'wet={np.mean(y):.3f}(n={len(y)}), '
          f'bound=±{b:.3f}{unit}, TOST p={p:.4f}, equiv={eq}')

wb = Workbook()
hdr_fill = PatternFill('solid', fgColor='1a4a2e')
hdr_font = Font(bold=True, color='FFFFFF', name='Arial', size=10)
val_font = Font(name='Arial', size=9)
alt_fill = PatternFill('solid', fgColor='eef5f0')
bas_fill = PatternFill('solid', fgColor='ffffff')
fail_fill= PatternFill('solid', fgColor='fde0e0')
thin = Side(style='thin', color='aaaaaa')
brd  = Border(left=thin, right=thin, top=thin, bottom=thin)
ctr  = Alignment(horizontal='center', vertical='center')

def wh(ws, hdrs):
    for ci, h in enumerate(hdrs, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font=hdr_font; cell.fill=hdr_fill
        cell.alignment=ctr; cell.border=brd
        ws.column_dimensions[get_column_letter(ci)].width=max(14,len(str(h))*0.9)

def sv(ws, ri, ci, v, fail=False):
    fill = fail_fill if fail else (alt_fill if ri%2==0 else bas_fill)
    cell = ws.cell(row=ri, column=ci, value=v)
    cell.font=val_font; cell.fill=fill
    cell.alignment=ctr; cell.border=brd
    if isinstance(v, float) and not np.isnan(v):
        cell.number_format = '0.0000' if abs(v)<1000 else '0.0'

ws1 = wb.active; ws1.title = 'Per-Well'
hdrs1 = ['Plate','Well','Condition','Pass','Reason',
          'n_total','n_valid','valid_frac','cell_plane_um',
          'cell_plane_frac','intra_focus_sd','sawtooth','rho',
          'fit_model','C0','A','A_gain','mu_max','lag','kd','td','fit_R2','obs_max']
wh(ws1, hdrs1)
for ri, r in enumerate(all_wells, 2):
    row = [r['plate'], r['well'], r['condition'],
           'YES' if r['pass'] else 'no', r['reason'],
           r['n_total'], r['n_valid'],
           round(r['valid_fraction'],3),
           round(r['cell_plane_um'],1) if not np.isnan(r['cell_plane_um']) else None,
           round(r['cell_plane_fraction'],3),
           round(r['intra_focus_sd'],2) if not np.isnan(r['intra_focus_sd']) else None,
           round(r['sawtooth'],4) if not np.isnan(r['sawtooth']) else None,
           round(r['rho'],3) if not np.isnan(r['rho']) else None,
           r['fit_name'],
           round(r['C0'],3) if not np.isnan(r['C0']) else None,
           round(r['A'],3) if not np.isnan(r['A']) else None,
           round(r['A_gain'],3) if not np.isnan(r['A_gain']) else None,
           round(r['mu_max'],4) if not np.isnan(r['mu_max']) else None,
           round(r['lag'],2) if not np.isnan(r['lag']) else None,
           round(r['kd'],5) if not np.isnan(r['kd']) else None,
           round(r['td'],2) if not np.isnan(r['td']) else None,
           round(r['fit_r2'],3) if not np.isnan(r['fit_r2']) else None,
           round(r['obs_max'],2) if not np.isnan(r['obs_max']) else None]
    for ci, v in enumerate(row, 1):
        sv(ws1, ri, ci, v, fail=not r['pass'])
ws1.freeze_panes = 'E2'

ws2 = wb.create_sheet('Condition Summary')
wh(ws2, ['Condition','n_pass','n_total',
          'A_mean','A_sd','mu_max_mean','mu_max_sd',
          'lag_mean','lag_sd','obs_max_mean','obs_max_sd'])
ri = 2
for cond in COND_LIST:
    pa  = [r for r in all_wells if r['condition']==cond and r['pass']]
    tot = sum(1 for r in all_wells if r['condition']==cond)
    def ms(key):
        vs = [r[key] for r in pa if not np.isnan(r[key])]
        if not vs: return None, None
        return round(np.mean(vs),3), (round(np.std(vs,ddof=1),3) if len(vs)>1 else None)
    Am,As = ms('A'); Mm,Ms = ms('mu_max'); Lm,Ls = ms('lag'); Om,Os = ms('obs_max')
    for ci,v in enumerate([cond,len(pa),tot,Am,As,Mm,Ms,Lm,Ls,Om,Os],1):
        sv(ws2, ri, ci, v)
    ri += 1
ws2.freeze_panes = 'A2'

for label, param in [('Pairwise_A','A'),('Pairwise_mu_max','mu_max')]:
    ws = wb.create_sheet(label)
    df = pairwise_mwu(all_wells, param)
    if df.empty:
        ws.cell(row=1,column=1,value='no pairs'); continue
    wh(ws, list(df.columns))
    for ri, row in enumerate(df.itertuples(index=False), 2):
        for ci, v in enumerate(row, 1):
            if isinstance(v, float): v = round(v,4) if not np.isnan(v) else None
            sv(ws, ri, ci, v)

ws_t = wb.create_sheet('TOST')
wh(ws_t, ['Parameter','n_dried','n_wet','mean_dried','mean_wet',
           'diff','bound','TOST_p','equivalent'])
tri = 2
for key, bound, unit in [('A', TOST_BOUND_A, 'pp'),
                         ('mu_max', None, 'rel')]:
    x = gv('Dried Fibrinogen', key)
    y = gv('Wet Fibrinogen', key)
    if not x or not y: continue
    b = bound if bound else TOST_BOUND_MU * np.mean(y)
    p, eq = tost(x, y, b)
    for ci, v in enumerate([key,len(x),len(y),
                             round(np.mean(x),4),round(np.mean(y),4),
                             round(np.mean(x)-np.mean(y),4),
                             round(b,4),round(p,4),
                             'YES' if eq else 'no'], 1):
        sv(ws_t, tri, ci, v)
    tri += 1

ws_m = wb.create_sheet('Methods')
ws_m.column_dimensions['A'].width = 110
mlines = [
    ('METHODS — Focus Filtering and Growth Model Fitting (v3)', True),
    ('', False),
    ('1. Focus-based cell-plane identification (per well)', True),
    ('   Kernel density estimation (Gaussian, bw=50 µm) on raw focus values.', False),
    ('   Peaks ≥ 3000 µm are candidate cell-plane modes (cells sit on COC foil).', False),
    ('   The candidate with the most timepoints within ±40 µm is selected.', False),
    ('   No minimum fraction required — even a minority of on-focus timepoints', False),
    ('   is sufficient, provided there are enough for fitting (see below).', False),
    ('   Wells with NO peak ≥ 3000 µm are excluded (device failure).', False),
    ('', False),
    ('2. Valid timepoint extraction', True),
    (f'   All timepoints within ±{FOCUS_TOL:.0f} µm of the cell-plane mode.', False),
    (f'   Inclusion requires: ≥ {MIN_VALID_PTS} valid timepoints AND', False),
    (f'   time span of valid points ≥ {MIN_TIME_SPAN:.0f} h.', False),
    ('   Sawtooth index and pre-peak Spearman ρ are computed as informative', False),
    ('   metrics but do NOT exclude wells.', False),
    ('', False),
    ('3. Growth model fitting (on valid timepoints only)', True),
    ('   Four candidate models tested:', False),
    ('     (a) Gompertz (Zwietering 1990) with baseline offset C0: 4 params', False),
    ('         (C0, A, mu_max, lambda)', False),
    ('     (b) Gompertz with offset + exponential post-peak decline: 6 params', False),
    ('         (C0, A, mu_max, lambda, kd, td)', False),
    ('     (c) 4-parameter logistic with offset: 4 params (C0, A, k, tmid)', False),
    ('     (d) Linear: 2 params (fallback for flat traces, e.g. COC)', False),
    ('   Model selection: AICc (AIC with the small-sample correction,', False),
    ('   K counting the residual variance). The simplest model within', False),
    ('   Delta-AICc < 4 is chosen (Burnham & Anderson 2002). AICc rather', False),
    ('   than AIC because the shortest included well has 19 valid points', False),
    ('   against up to 6 curve parameters. Fit-parameters are unified', False),
    ('   across model types (A, mu_max, lambda extracted consistently).', False),
    ('   No R²-based exclusion: flat low-confluence curves (COC) have', False),
    ('   structurally low R² but are valid biological outcomes.', False),
    ('', False),
    ('4. Visualisation', True),
    ('   Fit curves evaluated on t = 0…140 h grid (500 points).', False),
    (f'   Reported A is the fitted plateau (upper asymptote). For the linear', False),
    (f'   model, which has no asymptote, it is the fitted value at t = {T_REF:.0f} h,', False),
    ('   a fixed reference time, so that wells of different record length', False),
    ('   stay comparable. A_gain is the corresponding rise above C0.', False),
    ('   Thin lines: individual per-well fit curves.', False),
    ('   Bold line: mean of all per-well fit curves.', False),
    ('   Shaded band: mean ± 1.96 × SE of per-well fit curves (≈95% CI).', False),
    ('   All curves start at t = 0 by construction.', False),
    ('', False),
    ('5. Statistics', True),
    ('   Pairwise: Mann-Whitney-U + Bonferroni correction.', False),
    ('   Primary question (dried vs. wet equivalence): TOST', False),
    (f'   Equivalence bounds: A ± {TOST_BOUND_A} pp, µ_max ± {TOST_BOUND_MU*100:.0f}% relative.', False),
]
for i, (text, bold) in enumerate(mlines, 1):
    cell = ws_m.cell(row=i, column=1, value=text)
    cell.font = Font(name='Arial', size=10, bold=bold,
                     color='1a4a2e' if bold else '333333')
    if bold and text:
        cell.fill = PatternFill('solid', fgColor='eef5f0')

wb.save('fibrinogen_v3_results.xlsx')
print('  Excel:    fibrinogen_v3_results.xlsx')
print('\nFERTIG.')
